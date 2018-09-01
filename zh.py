import yaml
from openpyxl import *

from openpyxl.utils import get_column_letter
from openpyxl.utils import column_index_from_string
f=open('C:\\Users\\admin\\Seafile\\snmp\\report_enginx\\wg_day_report.yaml','r',encoding='utf-8')
config=yaml.load(f)
wb=load_workbook("C:\\Users\\admin\\Seafile\\snmp\\共济自定义报表程序\\"  + config['filename'])
for i in config['sheet']:
    for x in config['sheet'][i]:
        #print(x)
        if x=='sheet_name': pass
        else:
            #print( config['sheet'][i])
            if config['sheet'][i][x]['type']=='database':
                text1=''
                text2=''
                #print(config['sheet'][i][x]['content']['devid'])
                for index,item in enumerate(config['sheet'][i][x]['content']['devid']):
                    if config['sheet'][i][x]['content']['formula']=='lj':
                    #print(item)
                        text1+='JLJ({})+'.format(get_column_letter((index+1)))
                        text2+='{}={}|{}\n'.format(get_column_letter((index+1)),item['devid'],item['point_index'])
                    elif config['sheet'][i][x]['content']['formula']=='avg':
                    #print(item)
                        text1+='JAVG({})+'.format(get_column_letter((index+1)))
                        text2+='{}={}|{}\n'.format(get_column_letter((index+1)),item['devid'],item['point_index'])
                text1='DATABASE|'+text1[:-1]
                print('__________________________\n',text2,text1)
                print((config['sheet'][i][x]['index'])[:-2]+'3')
                wb[config['sheet'][i]['sheet_name']][(config['sheet'][i][x]['index'])[:-2]+'5'].value=text2
                wb[config['sheet'][i]['sheet_name']][(config['sheet'][i][x]['index'])[:-2]+'6'].value=text1
            elif config['sheet'][i][x]['type']=='formula':
                text1='TEMPLATE|' +str(config['sheet'][i][x]['content']['formula'])
                print(config['sheet'][i][x]['index']+'3')
                wb[config['sheet'][i]['sheet_name']][(config['sheet'][i][x]['index'])[:-2]+'6'].value=text1

wb.save('ss.xlsx')


    


