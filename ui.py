# -*- coding: utf_8 -*-

from datetime import date
import time
from os import path
from js2 import *
from img import *
#import base64
import logging 
import logging.handlers 
from openpyxl import load_workbook
from PyQt5.QtWidgets import (QTextEdit,QAction, qApp,QApplication,QMainWindow,QMessageBox,QDateTimeEdit,QComboBox,QFileDialog,QPushButton,QProgressBar,QGridLayout,QApplication, QCheckBox, QDialog,
        QDialogButtonBox, QFrame, QGroupBox, QLabel, QLineEdit, QListWidget,
        QTabWidget, QVBoxLayout, QWidget)
from PyQt5.QtCore import QCoreApplication,Qt,QDateTime,QDate,QTime  ,QThread,pyqtSignal,QFileInfo
from PyQt5.QtGui import QPixmap,QIcon,QDesktopServices
import configparser
from g_data import bl
 
#####################################
from ctypes import windll
windll.shell32.SetCurrentProcessExplicitAppUserModelID("报表程序")
###########################
blc=bl()   ##放了一些变量,用来作于全局变量导入

########################初始化log日志
LEVELS={'notset':logging.DEBUG,  
        'debug':logging.DEBUG,   
        'info':logging.INFO,   
        'warning':logging.WARNING,  
        'error':logging.ERROR,    
        'critical':logging.CRITICAL} 
  

LOG_FILENAME = 'test.log'  
LOG_BACKUPCOUNT = 5   
LOG_LEVEL = 'notset'      
def InitLog(file_name,logger):   
    LOG_FILENAME = file_name       
    handler = logging.handlers.RotatingFileHandler(LOG_FILENAME,maxBytes=10*1024*1024,backupCount=LOG_BACKUPCOUNT)    
    #handler = logging.FileHandler(LOG_FILENAME)  
    formatter = logging.Formatter("[ %(asctime)s ][ %(levelname)s ] %(message)s")   
    handler.setFormatter(formatter)       
    #logger = logging.getLogger()   
    logger.addHandler(handler)   
    logger.setLevel(LEVELS.get(LOG_LEVEL.lower()))  
    return logger  

logger2 = logging.getLogger('requestinfo')
logger2=InitLog('报表日志{}.log'.format(time.strftime('%Y%m%d',time.localtime(time.time()))),logger2)
##################
#两个函数，用来excel中列的转换字母转数字数字转字母，
#代替openpyxl中的默认函数，默认的用了numpy库，这个库效率高，但是太大了

def get_column_letter(n):
    x=26
    n-=1
    a=['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N',
     'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z']
    b=[]
    while True:
        s=n//x-1#商
        y=n%x#余数
        b=b+[y]
        if s+1<=0:
            break
        n=s
    b.reverse()
    strx=''
    for i in b:
        strx+=a[i]
    return strx
def column_index_from_string(x):
    x=x.upper()
    a=['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N',
     'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z']
    b=[]
    for index,item in enumerate(x):
        b.append(a.index(item))
    val=0
    lenb=len(b)

    for index,i in enumerate(b):
        val+=(i+1)*(26**(lenb-index-1))





####################多线程函数##########################################
class WorkThread(QThread):   
    sinOut = pyqtSignal(int)
    def __int__(self):  
        super(WorkThread,self).__init__()
    def init2(self,filename,stdate,enddate,gzh,sjl,cj,mode):
        self.filename=filename
        self.stdate=stdate
        self.enddate=enddate
        self.gzh=gzh
        self.sjl=sjl
        self.cj=cj
        self.mode=mode
    def run(self):
        self.main()
    def init(self,host='172.31.61.61',port=3307,user='gj',pwd='xbrother',db='historyver1'):  #初始化mysql
        if self.cj=='共济':
            try:
                from   mysql import MYSQL
                import  xb_1 as cv 
                self.cv=cv
                self.ms=MYSQL(host=host,port=port,user=user,pwd=pwd,db=db)
                self.flag=1
            except:
                self.flag=0
                self.sinOut.emit(2000)
        elif self.cj=='中联':
            try:
                from   mssql import MSSQL
                import zl_1 as cv
                self.cv=cv
                self.ms=MSSQL(host=host,port=port,user=user,pwd=pwd,db=db)
                self.flag=1
            except:
                self.flag=0
                self.sinOut.emit(2001)
    def finddate(self,ws,sjl,date):  ##在表格的某一列找到对应的日期
        flag=0
        for datelist in ws[sjl+str(int(self.gzh)+1)+':'+sjl+str(ws.max_row)]:
            timer=datelist[0].value
            try:
                if timer.year ==date.tm_year and timer.month==date.tm_mon and timer.day==date.tm_mday :
                    return datelist[0].row
                else :
                    flag+=1
                if flag == len(ws[sjl+'1'+':'+sjl+str(ws.max_row)]):
                    return False
            except:return False


    def main(self):  #主函数
        #print(self.filename)
        self.init(host=blc.host,port=int(blc.port),user=blc.usr,pwd=blc.passwd,db=blc.db)
        if  self.flag:
            wb=load_workbook(self.filename)
            xdate=0
            xled=0
            for i in wb.sheetnames:   #计算总共需要处理多少个单元格
                xdate+=len(wb[i][self.sjl+self.gzh:get_column_letter(wb[i].max_column)+self.gzh][0][1:])
            xdate=xdate*(self.enddate-self.stdate)/(60*60*24)   #计算总共需要处理多少个单元格  乘以需要处理的天数
            #print(xdate)
            row_x=wb[i].max_row
            for date1 in (range(int((self.enddate-self.stdate)/(60*60*24)))):
                if self.mode==0:
                    row_x=self.finddate(wb[i],self.sjl, time.localtime(self.stdate+(60*60*24*(date1))))
                elif self.mode==1:
                    row_x=row_x+1
                for index,i in enumerate(wb.sheetnames):   #依次处理的表格
                    if row_x:  #查找日期在不在
                        if self.mode==1:
                            dddddd=time.localtime(self.stdate+(60*60*24*(date1)))
                            wb[i][self.sjl+str(row_x)].value=date(dddddd.tm_year,dddddd.tm_mon,dddddd.tm_mday)
                        for x in wb[i][self.sjl+self.gzh:get_column_letter(wb[i].max_column)+self.gzh][0][1:]:  #每一列处理数据
                            cell_gz_value=wb[i][x.column+str(x.row)].value  #规则单元格数据
                            cell_cd_value=wb[i][x.column+str(x.row-1)].value   #测点单元格数据
                            try:
                                if cell_gz_value and cell_gz_value.startswith('TEMPLATE|'):   #如果字符以TEMPLATE开头
                                    tp=cell_gz_value.split('|')[1]   #查找括号内的内容
                                    tp=tp.replace('{}',str(row_x))   #吧{}替换成行数
                                    wb[i][x.column+str(row_x)].value=tp
                                elif cell_gz_value and cell_gz_value.startswith('DATABASE|'):  ##如果规则列存在数据
                                    cell_gz_value=cell_gz_value.split('|')[1]   #查找括号内的内容
                                    #print(cell_gz_value)
                                    exc=cell_gz_value   #获取规则 
                                    cd=cell_cd_value.split('\n')  #从文本中分离测点,添加到一个list中去
                                    cd = [zzj for zzj in cd if zzj !='']   #去除空的iten
                                    for bl in cd :  #处理每个测点的数据库的值
                                        bl=bl.split('=')   #分离测点,第一个元素将被命名为变量,第二个元素将作为测点ID用于数据库查找
                                        exc=exc.replace('({})'.format(bl[0]),'(locals()[\'{}\'])'.format(bl[0]))   #把规则中的变量替换成需要的全局变量
                                        stdate=self.stdate+60*60*24*(date1)
                                        enddate=self.stdate+60*60*24*(date1+1)
                                        locals()[bl[0]]=self.cv.sqldate(self.ms,logger2,stdate,enddate,bl[1])  #执行sql查询语句
                                    template='locals()[\'value\']= '+exc  #执行模板聚聚
                                    exec(template)  #执行模板聚聚
                                    wb[i][x.column+str(row_x)].value=locals()['value']   #得到结果填充到表格中去
                            except: pass
                            self.sinOut.emit(((xled)/xdate)*100)  #输出信号
                            xled+=1
                            #print(((xled)/xdate)*100)
                    else:
                        xled+=xdate/((self.enddate-self.stdate)/(60*60*24))
                        self.sinOut.emit(((xled)/xdate)*100)
            try:
                wb.save(self.filename)
            except:self.sinOut.emit(1000)
    
            self.sinOut.emit(100)
###############C初始化config.ini这个配置文件#########################################
config = configparser.ConfigParser()
configdict={'server':{'host':'192.168.1.93','port':'3307','user':'gj','passwd':'xbrother','db':'historyver1'},
        'sheet_day':{'sjl':'A',"gzh":'3',"cj":'1',"mode":'0',"filename":''}

        }

if path.exists("config.ini"):  #如果配置文件存在,导入配置文件
    config.read("config.ini")
else:

    print('初始化配置文件')   #如果配置文件不存在,从初始字典里面导入配置文件
    config.read_dict(configdict)
    config.write(open("config.ini","w"))
##########################################################
#################初始化图标文件，如果图标文件不纯在，从base64编码里面释放该图标文件##################
#import ico
#def  sfico(iconame):
#    if not os.path.exists(iconame):
#        with open(iconame, 'wb+') as tmp:   #临时文件用来保存jpg文件  
#            tmp.write(base64.b64decode(ico.img[iconame.split('.')[0]]))
#            tmp.close()
#icolist=  ['save1.ico','title.jpg']    
#for i in icolist:
#    sfico(i)

##################################################
class TabDialog(QDialog):           ##########主界面
    def __init__(self, parent=None):
        super(TabDialog, self).__init__(parent)
        #img = QPixmap(":/img/a.png")
        self.icon_title = QIcon()
        self.icon_title.addPixmap(QPixmap(":/img/title.ico"), QIcon.Normal, QIcon.Off)   #标题图表
        self.icon_save = QIcon()
        self.icon_save.addPixmap(QPixmap(":/img/save.ico"), QIcon.Normal, QIcon.Off)  #保存图表
        self.icon_help=QIcon()
        self.icon_help.addPixmap(QPixmap(":/img/help.ico"), QIcon.Normal, QIcon.Off)  # 帮助图表
        self.icon_aboat = QIcon()
        self.icon_aboat.addPixmap(QPixmap(":/img/about.ico"), QIcon.Normal, QIcon.Off)  #关于图表
        self.icon_github = QIcon()
        self.icon_github.addPixmap(QPixmap(":/img/github.ico"), QIcon.Normal, QIcon.Off)  #关于图表
        self.setWindowTitle('自定义报表程序')
        self.setGeometry(300,300,650,270) 
        #self.setFixedSize(self.width(), self.height())
        permissionsGroup = QGroupBox("服务器信息")
        permissionsLayout = QGridLayout()
        self.setWindowIcon(self.icon_title)
        #########################################
        qm=QMainWindow()
        exitAct = QAction(self.icon_save,"保存配置",qm)
        exitAct.setShortcut("Ctrl+S")
        exitAct.triggered.connect(self.saveconfig)
        self.toolbar = qm.addToolBar("save")
        exitAct1 = QAction(self.icon_help,"帮助",qm)
        exitAct1.setShortcut("Ctrl+H")
        exitAct1.triggered.connect(self.openhelp)
        exitAct2 = QAction(self.icon_aboat,"关于",qm)
        exitAct2.setShortcut("Ctrl+I")
        exitAct2.triggered.connect(self.openaboat)
        exitAct3 = QAction(self.icon_github,"查看源码",qm)
        exitAct3.setShortcut("Ctrl+M")
        exitAct3.triggered.connect(self.openaurl)
        self.toolbar.addAction(exitAct)
        self.toolbar.addAction(exitAct1)
        self.toolbar.addAction(exitAct2)
        self.toolbar.addAction(exitAct3)
        self.updatlabel=QLabel('',self.toolbar)
        self.updatlabel.setOpenExternalLinks(True)
        self.updatlabel.move(500,0)
        ###############第一行###############
        permissionsLayout.addWidget(QLabel('数据库',permissionsGroup),1,0 )
        self.le_host=QLineEdit('',permissionsGroup)
        self.le_host.editingFinished.connect(self.initserver)
        self.le_host.setInputMask('000.000.000.000')
        self.le_host.insert(config['server']['host'])
        self.le_host.setFixedWidth(100)
        #print((self.le_host.height(),self.le_host.width()))
        permissionsLayout.addWidget(QLabel('端口',permissionsGroup),1,2 )
        self.le_port=QLineEdit('',permissionsGroup)
        self.le_port.setInputMask('99999')
        self.le_port.insert(config['server']['port'])
        self.le_port.editingFinished.connect(self.initserver)
        self.le_port.setFixedWidth(40)
        permissionsLayout.addWidget(QLabel('用户名',permissionsGroup),1,4)
        self.le_usr=QLineEdit('',permissionsGroup)
        self.le_usr.insert(config['server']['user'])
        self.le_usr.editingFinished.connect(self.initserver)
        self.le_usr.setFixedWidth(40)
        permissionsLayout.addWidget(QLabel('密码',permissionsGroup),1,6 )
        self.le_passwd=QLineEdit('',permissionsGroup)
        self.le_passwd.insert(config['server']['passwd'])
        self.le_passwd.editingFinished.connect(self.initserver)
        self.le_passwd.setFixedWidth(100)
        self.le_passwd.setEchoMode(QLineEdit.PasswordEchoOnEdit)
        permissionsLayout.addWidget(QLabel('数据库名',permissionsGroup),1,8 )
        self.le_db=QLineEdit('',permissionsGroup)
        self.le_db.insert(config['server']['db'])
        self.le_db.setFixedWidth(80)
        self.le_db.editingFinished.connect(self.initserver)
        ##################################
        permissionsLayout.addWidget(self.le_host,1,1)
        permissionsLayout.addWidget(self.le_port,1,3)
        permissionsLayout.addWidget(self.le_usr,1,5)
        permissionsLayout.addWidget(self.le_passwd,1,7)
        permissionsLayout.addWidget(self.le_db,1,9)
        permissionsGroup.setLayout(permissionsLayout)
        mainLayout = QVBoxLayout()
        mainLayout.addWidget(qm)
        mainLayout.addWidget(permissionsGroup)
        mainLayout.addStretch(1)
        self.setLayout(mainLayout)
        tabWidget = QTabWidget()
        self.ts1=ri()
        self.ts2=zhou()
        self.ts3=yue()
        tabWidget.addTab(self.ts1, "日报")
        mainLayout.addWidget(tabWidget)
        tabWidget.addTab(self.ts2, "周报")
        mainLayout.addWidget(tabWidget)
        tabWidget.addTab(self.ts3, "月报")
        mainLayout.addWidget(tabWidget)
        self.checkvision()   #运行检查升级函数
    ###################检测版本########################################
    def  checkvision(self):
        import requests
        try:
            url='http://wechat.52hengshan.com/weixin/api/v1.0/appvison'
            s=requests.get(url=url).json()
        except:
            self.updatlabel.setText("<html><head/><body><p><span style=\" text-decoration: underline; color:#00ff00;\">无法联网查找最新版本</span></a></p></body></html>")
        else:
            if blc.vison==s['vison']:
                self.updatlabel.setText("<html><head/><body><p><span style=\" text-decoration: underline; color:#0000ff;\">已经是最新版本</span></a></p></body></html>")
            else:self.updatlabel.setText("<html><head/><body><p><a href=\"{}\"><span style=\" text-decoration: underline; color:#ff0000;\">有最新版本,点此升级</span></a></p></body></html>".format(s['url']))
    
    
    
    def initserver(self):   ##########从主界面获取的数据存入全局变量
        blc.host=self.le_host.text()
        blc.port=self.le_port.text()
        blc.usr=self.le_usr.text()
        blc.passwd=self.le_passwd.text()
        blc.db=self.le_db.text()
    def openaboat(self):   #######关于菜单的弹窗内容
        #QMessageBox.information(self, "关于", "V1.0\n业余作品\n作者:谭润芝\n电话:13267153721",QMessageBox.Yes)

        self.wid1 = QMainWindow() 
        self.wid1.setWindowTitle('关于')  
        self.wid1.setGeometry(300,300,300,300)
        self.wid1.setWindowIcon(self.icon_aboat)    
        reviewEdit = QTextEdit(self.wid1 )
        reviewEdit.setGeometry(0,0,300,300) 
        self.wid1.setFixedSize(self.wid1.width(), self.wid1.height())
        reviewEdit.setPlainText(blc.text_about)
        reviewEdit.setReadOnly(True)
        self.wid1.show()
    def openaurl(self):
        QDesktopServices.openUrl(QtCore.QUrl('https://github.com/trz0332'))
    def openhelp(self):          ##帮助菜单的弹窗内容
        #self.hide()
        if path.exists("help.txt"):    #判断help.txt是否存在，如果存在获取帮助文本里面的内容，填充到文本编辑控件显示
            with open('help.txt','r') as fb:
                blc.text_help=fb.read()
        else:
            with open('help.txt','a') as fb:   #如果帮助文本不存下，从默认的帮助变量获取文本，保存到help.txt
                fb.write(blc.text_help)
            #config.write(open("config.ini","w"))
        self.wid = QMainWindow() 
        self.wid.setWindowTitle('帮助')  #帮助窗口标题
        self.wid.setGeometry(300,300,600,400)   #帮助窗口大小
        self.wid.setWindowIcon(self.icon_help)    #帮助窗口图标
        self.reviewEdit = QTextEdit(self.wid )   #定义一个文本编辑控件
        self.reviewEdit.setGeometry(0,0,600,400)   #设置文本编辑控件大小
        self.wid.setFixedSize(self.wid.width(), self.wid.height())  #设置固定大小不允许改变
        self.reviewEdit.setPlainText(blc.text_help)   #帮助文本的内容填入到文本编辑框
        self.reviewEdit.setReadOnly(True)   #制度模式
        self.wid.show()   #展示窗口
    def saveconfig(self):  #保存配置到config.init
        config.set('server','host',self.le_host.text())
        config.set('server','port',self.le_port.text())
        config.set('server','user',self.le_usr.text())
        config.set('server','passwd',self.le_passwd.text())
        config.set('server','db',self.le_db.text())
        config.set('sheet_day','sjl',self.ts1.sjl.text())
        config.set('sheet_day','gzh',self.ts1.gzh.text())
        config.set('sheet_day','cj',str(self.ts1.jsComboBox.currentIndex()))
        config.set('sheet_day','mode',str(self.ts1.js2ComboBox.currentIndex()))
        config.set('sheet_day','mode',str(self.ts1.js2ComboBox.currentIndex()))
        config.set('sheet_day','filename',str(self.ts1.qtfile.text()))
        with open("config.ini","w") as fh:
            config.write(fh)

        QMessageBox.information(self, "信息", "配置文件保存成功",QMessageBox.Yes)
class ri(QWidget):    #日报table框组建
    def __init__(self,parent=None):
        super(ri, self).__init__(parent)
        self.workThread=WorkThread()
        self.workThread.sinOut.connect(self.setp)
        layout = QGridLayout()
        self.fileName1=''
        layout.addWidget(QLabel('报表文件',self),0,0 )
        self.qtfile=QLineEdit('',self)
        self.qtfile.insert(config['sheet_day']['filename'] if 'filename' in  config['sheet_day'].keys() else '')
        self.qbfile=QPushButton('选择文件',self)
        self.qbfile.clicked.connect(self.button_click)
        self.qbfile.setEnabled(True)
        layout.addWidget(self.qtfile,0,1,1,6)
        layout.addWidget(self.qbfile,0,7)
        ######################################################################
        layout.addWidget(QLabel('规则行',self) ,1,0)
        self.gzh=QLineEdit('',self)
        self.gzh.insert(config['sheet_day']['gzh'])
        self.gzh.setFixedWidth(40)
        layout.addWidget(QLabel('时间列',self) ,1,2)
        self.sjl=QLineEdit('',self)
        self.sjl.insert(config['sheet_day']['sjl'])
        self.sjl.setFixedWidth(40)
        layout.addWidget(QLabel('BMS厂家',self) ,1,4)
        jslist=['共济','中联']
        self.jsComboBox=QComboBox(self)   #下拉菜单
        for i in jslist:
            self.jsComboBox.addItem(i)
        #print(help(self.jsComboBox))
        self.jsComboBox.setCurrentIndex(int(config['sheet_day']['cj']))
        layout.addWidget(QLabel('模式',self),1,6)
        jslist2=['0--寻找日期','1--最后填充']
        self.js2ComboBox=QComboBox(self)   #下拉菜单
        for i in jslist2:
            self.js2ComboBox.addItem(i)
        self.js2ComboBox.setCurrentIndex(int(config['sheet_day']['mode']))
        layout.addWidget(self.gzh,1,1)
        layout.addWidget(self.sjl,1,3)
        layout.addWidget(self.jsComboBox,1,5)
        layout.addWidget(self.js2ComboBox,1,7)
        ###################################################
        layout.addWidget(QLabel('起始时间',self),2,0)
        self.dt1 = QDateTimeEdit(QDate.currentDate(),self) # 创建日期，并初始值  
        self.dt1.setDisplayFormat('yyyy-MM-dd')
        #self.dt1.setMinimumDate(QDate.currentDate().addDays(-90)) # 限定时间最小值，当前时间-365天  
        self.dt1.setMaximumDate(QDate.currentDate().addDays(1)) # 限定时间最大值，当前时间+365天  
        self.dt1.setCalendarPopup(True) # 允许弹出窗口选择日期，setMinimumDate()的限定对这个窗口也有效  
        # #############结束时间选择 
        layout.addWidget(QLabel('结束时间',self),2,4)
        self.dt2 = QDateTimeEdit(QDate.currentDate(),self) # 创建日期，并初始值  
        self.dt2.setDate(QDate.currentDate())  
        #self.dt2.setMinimumDate(QDate.currentDate().addDays(-90)) # 限定时间最小值，当前时间-365天  
        self.dt2.setMaximumDate(QDate.currentDate().addDays(1)) # 限定时间最大值，当前时间+365天  
        self.dt2.setCalendarPopup(True) # 允许弹出窗口选择日期，setMinimumDate()的限定对这个窗口也有效  

        ###########查询按钮###########
        self.qtb1=QPushButton('查询',self)
        #qtb1.setGeometry(30, 30, 30, 60)

        self.qtb1.clicked.connect(self.work)
        
        layout.addWidget(self.dt1,2,1,1,2)
        layout.addWidget(self.dt2,2,5,1,2)
        layout.addWidget(self.qtb1,2,7)


        layout.addWidget(QLabel('进度',self),3,0 )
        self.pbar = QProgressBar(self)
        self.pbar.setFixedWidth(560)
        layout.addWidget(self.pbar,3,1,1,8 )
        self.setLayout(layout)
    def setp(self,num):
        if num<100:
            self.pbar.setValue(num)
        elif num==1000:
            self.qtb1.setEnabled(True)
            self.pbar.setValue(0)
            self.workThread.quit()
            QMessageBox.information(self, "错误", "文件保存失败,查看文件是否被打开",QMessageBox.Yes)
        elif num==2000:
            self.qtb1.setEnabled(True)
            self.pbar.setValue(0)
            #self.workThread.quit()
            QMessageBox.information(self, "错误", "无法导入pymysql依赖包",QMessageBox.Yes)
        elif num==2001:
            self.qtb1.setEnabled(True)
            self.pbar.setValue(0)
            #self.workThread.quit()
            QMessageBox.information(self, "错误", "无法导入msmysql依赖包",QMessageBox.Yes)

        elif num >=100 and num <101 :
            self.qtb1.setEnabled(True)
            self.pbar.setValue(num)
            QMessageBox.information(self, "提示", "报表导出成功",QMessageBox.Yes)
    def over():
        self.qtb1.setEnabled(True)
    def work(self):

        self.t1=time.mktime(time.strptime(self.dt1.date().toString(Qt.ISODate),"%Y-%m-%d"))
        self.t2=time.mktime(time.strptime(self.dt2.date().toString(Qt.ISODate),"%Y-%m-%d"))+60*60*24
        gzh=self.gzh.text()
        sjl=self.sjl.text()
        cj=self.jsComboBox.currentText()
        self.fileName1=self.qtfile.text()
        mode=int(self.js2ComboBox.currentText().split('--')[0])
        if self.qtfile.text()=='':QMessageBox.information(self, "提示", "没有导入日报文件",QMessageBox.No)
        else:
            self.qtb1.setEnabled(False)
            self.workThread.init2(self.fileName1,self.t1,self.t2,gzh,sjl,cj,mode)
            self.workThread.start()              #计时开始  
            #self.workThread.trigger.connect(over)   #当获得循环完毕的信号时，停止计数 
    
    def button_click(self): 
        # absolute_path is a QString object  
        self.fileName1, self.filetype = QFileDialog.getOpenFileName(self,
                                            "选取文件",  
                                            "",  
                                            "xlsx (*.xlsx)")   #设置文件扩展名过滤,注意用双分号间隔  

        self.qtfile.clear()
        self.qtfile.insert(self.fileName1)
class zhou(QWidget):
    def __init__(self,parent=None):
        super(zhou, self).__init__(parent)
        layout = QVBoxLayout()
        topLabel = QLabel("还没做这个功能")
        layout.addWidget(topLabel)
        self.setLayout(layout)
class yue(QWidget):
    def __init__(self, parent=None):
        super(yue, self).__init__(parent)
        layout = QVBoxLayout()
        topLabel = QLabel("还没做这个功能")
        layout.addWidget(topLabel)
        self.setLayout(layout)

if __name__ == '__main__':

    import sys

    app = QApplication(sys.argv)


    tabdialog = TabDialog()
    tabdialog.show()
    sys.exit(app.exec_())



