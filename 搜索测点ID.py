from openpyxl import *
from   mysql import MYSQL
ms=MYSQL(host='192.168.1.93',port=10003,user='gj',pwd='xbrother',db='historyver1')
wb=load_workbook('2期新版PUEExcel 工作表 2018年6月份.xlsx') 
ws=wb['环境']  
for i in ws['B8:G8'][0]:
	print(i)
	sql1='SELECT device_no from t_device_info where device_name like \'{}\''.format(i.value)
	devid=ms.ExecQuery(sql1)
	if  devid:
		devid=devid[0][0]
		sql2="SELECT tag_no from t_tag_info where device_no like '{}' and tag_name like '{}'".format(devid,ws[i.column+'9'].value)
		print(sql2)
		tag_no=ms.ExecQuery(sql2)[0][0]
		print(tag_no)
		ws[i.column+'10'].value=tag_no
wb.save('2期新版PUEExcel 工作表 2018年6月份.xlsx')
