﻿

from PyQt5.QtCore import QThread,pyqtSignal
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle,Alignment,Border,Side
import time
import datetime
from .uilog import logger2 as log
from .js2 import *



#两个函数，用来excel中列的转换字母转数字数字转字母，
#代替openpyxl中的默认函数，默认的用了numpy库，这个库效率高，但是太大了
##################################初始化EXCEL黑色边框
align = Alignment(horizontal='left',vertical='center',wrap_text=True)
border = Border(left=Side(border_style='thin',color='000000'),
right=Side(border_style='thin',color='000000'),
top=Side(border_style='thin',color='000000'),
bottom=Side(border_style='thin',color='000000'))


#####################################
def get_column_letter(n):
    x=26
    n-=1
    a=['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N',
     'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z']
    b=[]
    while True:
        s=n//x-1#商
        y=n%x#余数
        b=b+[y]
        if s+1<=0:
            break
        n=s
    b.reverse()
    strx=''
    for i in b:
        strx+=a[i]
    return strx
def column_index_from_string(x):
    x=x.upper()
    a=['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N',
     'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z']
    b=[]
    for index,item in enumerate(x):
        b.append(a.index(item))
    val=0
    lenb=len(b)
    for index,i in enumerate(b):
        val+=(i+1)*(26**(lenb-index-1))
    return val
####################多线程函数##########################################
class WorkThread(QThread):   
    sinOut = pyqtSignal(int)
    def __int__(self):  
        super(WorkThread,self).__init__()
    def init2(self,filename,stdate,enddate,gzh,sjl,cj,mode,sheetnames,save_mode):
        self.filename=filename
        self.stdate=stdate
        self.enddate=enddate
        if gzh.isalpha():
            self.flag_hl=0   #如果规则是字母，则时间配列行
        else:self.flag_hl=1   #如果规则是数字，则时间在列上面
        self.gzh=gzh
        self.sjl=sjl
        self.cj=cj
        self.mode=mode
        self.save_mode=save_mode
        self.sheetnames=sheetnames
        #print(self.sheetnames)
    def init1(self,host,port,user,pwd,db):
        self.host=host
        self.port=port
        self.user=user
        self.pwd=pwd
        self.db=db
    def run(self):
        self.main()
    def init(self):  #初始化mysql
        #print(self.cj)
        if self.cj=='共济V7':
            try:
                from   .mysql import MYSQL
                from   . import xb_1 as cv 
                self.cv=cv
                self.ms=MYSQL(host=self.host,port=self.port,user=self.user,pwd=self.pwd,db=self.db)
                self.flag=1
            except:
                self.flag=0
                self.sinOut.emit(2000)
        elif self.cj=='中联S80':
            try:
                from   .mssql import MSSQL
                from   . import zl_1 as cv
                self.cv=cv
                self.ms=MSSQL(host=self.host,port=self.port,user=self.user,pwd=self.pwd,db=self.db)
                self.flag=1
            except:
                self.flag=0
                self.sinOut.emit(2001)
        elif self.cj=='栅格':
            try:
                from   .mysql import MYSQL
                from   . import sg_1 as cv
                self.cv=cv
                self.ms=MYSQL(host=self.host,port=self.port,user=self.user,pwd=self.pwd,db=self.db)
                self.flag=1
            except:
                self.flag=0
                self.sinOut.emit(2000)
        elif self.cj=='共济KE':
            try:
                import pyssdb
                from   . import ke_1 as cv
                self.cv=cv
                self.ms=pyssdb.Client(host=self.host, port=self.port)
                self.flag=1
            except Exception as e:
                print(e)
                self.flag=0
                self.sinOut.emit(2000)
        elif self.cj=='中联S90':
            try:
                import pymongo
                from   . import zl_2 as cv
                self.cv=cv
                self.ms=pymongo.MongoClient("mongodb://{}:{}/".format(self.host,self.port))['mon_data']
                self.flag=1
            except:
                self.flag=0
                self.sinOut.emit(2000)
        elif self.cj=='栅格clickhouse':
            try:
                from   .clickhouse import clickhouse
                from   . import sg_2 as cv
                self.cv=cv
                self.ms=clickhouse(host=self.host,port=self.port,user=self.user,pwd=self.pwd,db=self.db)
                self.flag=1
            except Exception as e :
                self.flag=0
                log.error(e)
                self.sinOut.emit(2000)
    def finddate(self,ws,date,fg=1):  ##在表格的某一列找到对应的日期
        flag=0
        if fg==1:
            for datelist in ws[self.sjl+str(int(self.gzh)+1)+':'+self.sjl+str(ws.max_row)]:
                timer=datelist[0].value
                try:
                    if timer.year ==date.tm_year and timer.month==date.tm_mon and timer.day==date.tm_mday :
                        return datelist[0].row
                    else :
                        flag+=1
                    if flag == len(ws[self.sjl+'1'+':'+self.sjl+str(ws.max_row)]):
                        log.info('没有找到日期{}---{}'.format(date,timer))
                        return False
                except:
                    flag=-1
                    log.info('查找日期函数出错，检查日期格式{}---{}'.format(date,timer))
                    #return False
        else:
            #fdt=get_column_letter(column_index_from_string(self.gzh)+1)  #计算时间的列号
            #print(fdt+str(int(self.sjl)),fdt+str(ws.max_column))
            log.info(ws[get_column_letter(column_index_from_string(self.gzh)+1)+str(int(self.sjl))+':'+get_column_letter(ws.max_column)+str(int(self.sjl))][0])
            #print(len(list(ws[get_column_letter(column_index_from_string(self.gzh)+1)+str(int(self.sjl))+':'+get_column_letter(ws.max_column)+str(int(self.sjl))])[0]))
            for datelist in list(ws[get_column_letter(column_index_from_string(self.gzh)+1)+str(int(self.sjl))+':'+get_column_letter(ws.max_column)+str(int(self.sjl))])[0]: #计算时间行的数据
                #print(datelist)
                timer=datelist.value
                #print(timer)
                try:
                    if timer.year ==date.tm_year and timer.month==date.tm_mon and timer.day==date.tm_mday :
                        return datelist.column
                    else :
                        flag+=1
                    if flag == len(ws[get_column_letter(column_index_from_string(self.gzh)+1)+str(int(self.sjl))+':'+get_column_letter(ws.max_column)+str(int(self.sjl))][0]):
                        log.info('没有找到日期{}---{}'.format(date,timer))
                        return False
                except Exception as e:
                    flag=-1
                    log.error('查找日期函数出错，检查日期格式{}---{}\n'.format(date,timer,e))
        if flag==-1:
            return False
                    #return False



    def main(self):  #主函数，
        self.init()
        #date_style = NamedStyle(name='datetime', number_format='YYYY-MM-DD')
        align = Alignment(horizontal='right', vertical='center')
        if  self.flag:
            wb=load_workbook(self.filename)
            xdate=0
            xled=0
            if self.flag_hl==1:    #处理时间在列上面的分支
                for i in self.sheetnames:   #计算总共需要处理多少个单元格
                    xdate+=len(wb[i][self.sjl+self.gzh:get_column_letter(wb[i].max_column)+self.gzh][0][1:])
                xdate=xdate*(self.enddate-self.stdate)/(60*60*24)   #计算总共需要处理多少个单元格  乘以需要处理的天数
                #print(xdate)
                for date1 in (range(int((self.enddate-self.stdate)/(60*60*24)))):
                    for index,sheet_name in enumerate(self.sheetnames):   #依次处理的表格
                        row_x=wb[sheet_name].max_row
                        if self.mode==0:
                            row_x=self.finddate(wb[sheet_name], time.localtime(self.stdate+(60*60*24*(date1))),self.flag_hl)
                        elif self.mode==1:
                            row_x=row_x+1
                        #print(sheet_name)  #打印表格名字
                        if row_x:  #查找日期在不在
                            if self.mode==1:
                                datetime.date
                                dddddd=time.localtime(self.stdate+(60*60*24*(date1)))
                                wb[sheet_name][self.sjl+str(row_x)].value=datetime.date(dddddd.tm_year,dddddd.tm_mon,dddddd.tm_mday)
                                #wb[i][self.sjl+str(row_x)].value=time.strftime("%Y-%m-%d", dddddd)
                                wb[sheet_name][self.sjl+str(row_x)].number_format = 'YYYY-MM-DD'
                                wb[sheet_name][self.sjl+str(row_x)].border=border   #得到结果填充到表格中去
                                wb[sheet_name][self.sjl+str(row_x)].alignment=align
                            #print(self.sjl+self.gzh,get_column_letter(wb[sheet_name].max_column)+self.gzh)
                            for x in wb[sheet_name][self.sjl+self.gzh:get_column_letter(wb[sheet_name].max_column)+self.gzh][0][1:]:  #每一列处理数据   
                                #print(x.column_letter,str(x.row))   #打印列号和行号
                                #print(x.column_letter)
                                #print(help(x))
                                try:
                                    cell_gz_value=wb[sheet_name][x.column_letter+str(x.row)].value  #规则单元格数据
                                    dtt=time.strftime("%Y-%m-%d",time.localtime(self.stdate+(60*60*24*date1)))
                                    if cell_gz_value and cell_gz_value.startswith('TEMPLATE|'):   #如果字符以TEMPLATE开头
                                        tp=cell_gz_value.split('|')[1]   #查找括号内的内容
                                        tp=tp.replace('{}',str(row_x))   #吧{}替换成行数
                                        wb[sheet_name][x.column_letter+str(row_x)].value=tp
                                        wb[sheet_name][x.column_letter+str(row_x)].border=border   #得到结果填充到表格中去
                                        wb[sheet_name][x.column_letter+str(row_x)].alignment=align   #得到结果填充到表格中去
                                    elif cell_gz_value and cell_gz_value.startswith('DATABASE|'):  ##如果规则列存在数据
                                        cell_gz_value=cell_gz_value.split('|')[1]   #查找括号内的内容
                                        cell_cd_value=wb[sheet_name][x.column_letter+str(x.row-1)].value.replace(' ','')   #测点单元格数据
                                        #print(cell_gz_value)
                                        exc=cell_gz_value   #获取规则 
                                        #print(cell_gz_value,cell_cd_value)
                                        cd=cell_cd_value.split('\n')  #从文本中分离测点,添加到一个list中去
                                        cd = [zzj for zzj in cd if zzj !='']   #去除空的iten
                                        for bl in cd :  #处理每个测点的数据库的值
                                            bl=bl.split('=')   #分离测点,第一个元素将被命名为变量,第二个元素将作为测点ID用于数据库查找
                                            exc=exc.replace('({})'.format(bl[0]),'(locals()[\'{}\'])'.format(bl[0]))   #把规则中的变量替换成需要的全局变量
                                            stdate=self.stdate+60*60*24*(date1)
                                            enddate=self.stdate+60*60*24*(date1+1)
                                            locals()[bl[0]]=self.cv.sqldate(self.ms,stdate,enddate,bl[1])  #执行sql查询语句
                                        template='locals()[\'value\']= '+exc  #执行模板聚he
                                        exec(template)  #执行模板聚
                                        wb[sheet_name][x.column_letter+str(row_x)].value=locals()['value']   #得到结果填充到表格中去
                                        wb[sheet_name][x.column_letter+str(row_x)].border=border   #得到结果填充到表格中去
                                        wb[sheet_name][x.column_letter+str(row_x)].alignment=align   #得到结果填充到表格中去
                                    log.info('处理完成{}日{}数据,规则{}，公式{}'.format(dtt,wb[sheet_name][x.column_letter+str(x.row-2)].value,cell_gz_value,cell_cd_value))
                                except  Exception as e: 
                                    log.error((e))
                                self.sinOut.emit(int(((xled)/xdate)*100)) #输出信号
                                xled+=1
                                
                                #print(((xled)/xdate)*100)
                        else:
                            xled+=xdate/((self.enddate-self.stdate)/(60*60*24))
                            if ((xled)/xdate)*100<100:
                                self.sinOut.emit(int(((xled)/xdate)*100))
            else:       #处理时间在行上面的分支
                for i in self.sheetnames:   #计算总共需要处理多少个单元格
                    xdate+=len(wb[i][self.gzh+str(int(self.sjl)+1):self.gzh+str(wb[i].max_row)])
                    #print(self.gzh+str(int(self.sjl)),self.gzh+str(wb[i].max_row))
                xdate=xdate*(self.enddate-self.stdate)/(60*60*24)   #计算总共需要处理多少个单元格  乘以需要处理的天数
                for date1 in (range(int((self.enddate-self.stdate)/(60*60*24)))):
                    for index,sheet_name in enumerate(self.sheetnames):   #依次处理的表格
                        row_x=wb[sheet_name].max_column
                        if self.mode==0:
                            row_x=self.finddate(wb[sheet_name], time.localtime(self.stdate+(60*60*24*(date1))),self.flag_hl)
                            print(row_x,'s')
                        elif self.mode==1:
                            row_x=row_x+1
                        #print(sheet_name)  #打印表格名字
                        if row_x:  
                            if self.mode==1:  #如果是最后填充，则先填充好日期头
                                datetime.date
                                dddddd=time.localtime(self.stdate+(60*60*24*(date1)))
                                wb[sheet_name][get_column_letter(row_x)+self.sjl].value=datetime.date(dddddd.tm_year,dddddd.tm_mon,dddddd.tm_mday)
                                #wb[i][self.sjl+str(row_x)].value=time.strftime("%Y-%m-%d", dddddd)
                                wb[sheet_name][get_column_letter(row_x)+self.sjl].number_format = 'YYYY-MM-DD'
                                wb[sheet_name][get_column_letter(row_x)+self.sjl].border=border
                                wb[sheet_name][get_column_letter(row_x)+self.sjl].alignment=align

                            #print(self.gzh+str(int(self.sjl)+1),self.gzh+str(wb[i].max_row))
                            for x in wb[sheet_name][self.gzh+str(int(self.sjl)+1):self.gzh+str(wb[sheet_name].max_row)]:  #遍历所有规则 
                                x=x[0]
                                #print(x.column_letter,x.row)   #打印列号和行号
                                try:
                                    cell_gz_value=x.value  #规则单元格数据
                                    dtt=time.strftime("%Y-%m-%d",time.localtime(self.stdate+(60*60*24*date1)))
                                    
                                    if cell_gz_value and cell_gz_value.startswith('TEMPLATE|'):   #如果字符以TEMPLATE开头
                                        tp=cell_gz_value.split('|')[1]   #查找括号内的内容
                                        #print(tp)
                                        tp=tp.replace('{}',get_column_letter(row_x))   #吧{}替换成列数
                                        wb[sheet_name][get_column_letter(row_x)+str(x.row)].value=tp    #将列号替换到{}中并填充
                                        wb[sheet_name][get_column_letter(row_x)+str(x.row)].border=border
                                        wb[sheet_name][get_column_letter(row_x)+str(x.row)].alignment =align
                                    elif cell_gz_value and cell_gz_value.startswith('DATABASE|'):  ##如果规则列存在数据
                                        cell_gz_value=cell_gz_value.split('|')[1]   #查找括号内的内容
                                        #print(cell_gz_value)
                                        exc=cell_gz_value   #获取规则 
                                        #print(cell_gz_value,cell_cd_value)
                                        cell_cd_value=wb[sheet_name][get_column_letter(x.column-1)+str(x.row)].value.replace(' ','')   #测点单元格数据
                                        cd=cell_cd_value.split('\n')  #从文本中分离测点,添加到一个list中去
                                        cd = [zzj for zzj in cd if zzj !='']   #去除空的iten
                                        for bl in cd :  #处理每个测点的数据库的值
                                            bl=bl.split('=')   #分离测点,第一个元素将被命名为变量,第二个元素将作为测点ID用于数据库查找
                                            exc=exc.replace('({})'.format(bl[0]),'(locals()[\'{}\'])'.format(bl[0]))   #把规则中的变量替换成需要的全局变量
                                            stdate=self.stdate+60*60*24*(date1)
                                            enddate=self.stdate+60*60*24*(date1+1)
                                            locals()[bl[0]]=self.cv.sqldate(self.ms,stdate,enddate,bl[1])  #执行sql查询语句
                                        template='locals()[\'value\']= '+exc  #执行模板聚he
                                        exec(template)  #执行模板聚
                                        #print(get_column_letter(row_x),x.row)
                                        wb[sheet_name][get_column_letter(row_x)+str(x.row)].value=locals()['value']   #得到结果填充到表格中去
                                        wb[sheet_name][get_column_letter(row_x)+str(x.row)].border=border
                                        wb[sheet_name][get_column_letter(row_x)+str(x.row)].alignment =align
                                    log.info('处理完成{}日{}数据,规则{}，公式{}'.format(dtt,wb[sheet_name][get_column_letter(x.column-2)+str(x.row)].value,cell_gz_value,cell_cd_value))
                                except  Exception as e: 
                                    log.error((e))
                                self.sinOut.emit(int(((xled)/xdate)*100))  #输出信号
                                xled+=1

                                #
                        else:
                            xled+=xdate/((self.enddate-self.stdate)/(60*60*24))
                            if ((xled)/xdate)*100<100:
                                self.sinOut.emit(int(((xled)/xdate)*100))
                        
                        #print(((xled)/xdate)*100)

            try:
                if self.save_mode:

                    wb.save(self.filename.split('.')[0]+'_'+time.strftime("%Y%m%d%H%M%S", time.localtime())+'.xlsx')

                else:
                    wb.save(self.filename)
                wb.close()
            except:self.sinOut.emit(1000)
    
            self.sinOut.emit(100)