from PyQt5.QtCore import QThread,pyqtSignal
from openpyxl import load_workbook
import time
from .uilog import logger2 as log
from .js2 import *
#两个函数，用来excel中列的转换字母转数字数字转字母，
#代替openpyxl中的默认函数，默认的用了numpy库，这个库效率高，但是太大了
def get_column_letter(n):
    x=26
    n-=1
    a=['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N',
     'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z']
    b=[]
    while True:
        s=n//x-1#商
        y=n%x#余数
        b=b+[y]
        if s+1<=0:
            break
        n=s
    b.reverse()
    strx=''
    for i in b:
        strx+=a[i]
    return strx
def column_index_from_string(x):
    x=x.upper()
    a=['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N',
     'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z']
    b=[]
    for index,item in enumerate(x):
        b.append(a.index(item))
    val=0
    lenb=len(b)
    for index,i in enumerate(b):
        val+=(i+1)*(26**(lenb-index-1))
####################多线程函数##########################################
class WorkThread(QThread):   
    sinOut = pyqtSignal(int)
    def __int__(self):  
        super(WorkThread,self).__init__()
    def init2(self,filename,stdate,enddate,gzh,sjl,cj,mode):
        self.filename=filename
        self.stdate=stdate
        self.enddate=enddate
        self.gzh=gzh
        self.sjl=sjl
        self.cj=cj
        self.mode=mode
    def init1(self,host,port,user,pwd,db):
        self.host=host
        self.port=port
        self.user=user
        self.pwd=pwd
        self.db=db
    def run(self):
        self.main()
    def init(self):  #初始化mysql
        if self.cj=='共济':
            try:
                from   mysql import MYSQL
                import  xb_1 as cv 
                self.cv=cv
                self.ms=MYSQL(host=self.host,port=self.port,user=self.user,pwd=self.pwd,db=self.db)
                self.flag=1
            except:
                self.flag=0
                self.sinOut.emit(2000)
        elif self.cj=='中联':
            try:
                from   mssql import MSSQL
                import zl_1 as cv
                self.cv=cv
                self.ms=MSSQL(host=self.host,port=self.port,user=self.user,pwd=self.pwd,db=self.db)
                self.flag=1
            except:
                self.flag=0
                self.sinOut.emit(2001)
        elif self.cj=='栅格':
            try:
                from   mysql import MYSQL
                import sg_1 as cv
                self.cv=cv
                self.ms=MYSQL(host=self.host,port=self.port,user=self.user,pwd=self.pwd,db=self.db)
                self.flag=1
            except:
                self.flag=0
                self.sinOut.emit(2000)
    def finddate(self,ws,sjl,date):  ##在表格的某一列找到对应的日期
        flag=0
        for datelist in ws[sjl+str(int(self.gzh)+1)+':'+sjl+str(ws.max_row)]:
            timer=datelist[0].value
            try:
                if timer.year ==date.tm_year and timer.month==date.tm_mon and timer.day==date.tm_mday :
                    return datelist[0].row
                else :
                    flag+=1
                if flag == len(ws[sjl+'1'+':'+sjl+str(ws.max_row)]):
                    return False
            except:return False


    def main(self):  #主函数，
        self.init()
        if  self.flag:
            wb=load_workbook(self.filename)
            xdate=0
            xled=0
            for i in wb.sheetnames:   #计算总共需要处理多少个单元格
                xdate+=len(wb[i][self.sjl+self.gzh:get_column_letter(wb[i].max_column)+self.gzh][0][1:])
            xdate=xdate*(self.enddate-self.stdate)/(60*60*24)   #计算总共需要处理多少个单元格  乘以需要处理的天数
            #print(xdate)
            row_x=wb[i].max_row
            for date1 in (range(int((self.enddate-self.stdate)/(60*60*24)))):
                if self.mode==0:
                    row_x=self.finddate(wb[i],self.sjl, time.localtime(self.stdate+(60*60*24*(date1))))
                elif self.mode==1:
                    row_x=row_x+1
                for index,i in enumerate(wb.sheetnames):   #依次处理的表格
                    if row_x:  #查找日期在不在
                        if self.mode==1:
                            dddddd=time.localtime(self.stdate+(60*60*24*(date1)))
                            wb[i][self.sjl+str(row_x)].value=date(dddddd.tm_year,dddddd.tm_mon,dddddd.tm_mday)
                        for x in wb[i][self.sjl+self.gzh:get_column_letter(wb[i].max_column)+self.gzh][0][1:]:  #每一列处理数据      
                            cell_gz_value=wb[i][x.column+str(x.row)].value  #规则单元格数据
                            cell_cd_value=wb[i][x.column+str(x.row-1)].value   #测点单元格数据
                            dtt=time.strftime("%Y-%m-%d",time.localtime(self.stdate+(60*60*24*date1)))
                            log.info('开始处理{}日{}数据,规则{}，公式{}'.format(dtt,wb[i][x.column+str(x.row-2)].value,cell_gz_value,cell_cd_value))
                            try:
                                if cell_gz_value and cell_gz_value.startswith('TEMPLATE|'):   #如果字符以TEMPLATE开头
                                    tp=cell_gz_value.split('|')[1]   #查找括号内的内容
                                    tp=tp.replace('{}',str(row_x))   #吧{}替换成行数
                                    wb[i][x.column+str(row_x)].value=tp
                                elif cell_gz_value and cell_gz_value.startswith('DATABASE|'):  ##如果规则列存在数据
                                    cell_gz_value=cell_gz_value.split('|')[1]   #查找括号内的内容
                                    #print(cell_gz_value)
                                    exc=cell_gz_value   #获取规则 
                                    cd=cell_cd_value.split('\n')  #从文本中分离测点,添加到一个list中去
                                    cd = [zzj for zzj in cd if zzj !='']   #去除空的iten
                                    for bl in cd :  #处理每个测点的数据库的值
                                        bl=bl.split('=')   #分离测点,第一个元素将被命名为变量,第二个元素将作为测点ID用于数据库查找
                                        exc=exc.replace('({})'.format(bl[0]),'(locals()[\'{}\'])'.format(bl[0]))   #把规则中的变量替换成需要的全局变量
                                        stdate=self.stdate+60*60*24*(date1)
                                        enddate=self.stdate+60*60*24*(date1+1)
                                        locals()[bl[0]]=self.cv.sqldate(self.ms,stdate,enddate,bl[1])  #执行sql查询语句
                                    template='locals()[\'value\']= '+exc  #执行模板聚he
                                    exec(template)  #执行模板聚
                                    wb[i][x.column+str(row_x)].value=locals()['value']   #得到结果填充到表格中去
                            except: pass
                            self.sinOut.emit(((xled)/xdate)*100)  #输出信号
                            xled+=1
                            #print(((xled)/xdate)*100)
                    else:
                        xled+=xdate/((self.enddate-self.stdate)/(60*60*24))
                        if ((xled)/xdate)*100<100:
                            self.sinOut.emit(((xled)/xdate)*100)
            try:
                wb.save(self.filename)
            except:self.sinOut.emit(1000)
    
            self.sinOut.emit(100)